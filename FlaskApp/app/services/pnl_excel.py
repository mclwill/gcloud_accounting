# FlaskApp/app/services/pnl_excel.py
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter


def build_pnl_workbook(pnl: Dict[str, Any], entity_name: str, subtitle: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit and Loss"

    bold = Font(bold=True)
    money_fmt = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'


    # Title
    ws["A1"] = entity_name
    ws["A1"].font = bold
    ws["A2"] = "Profit and Loss"
    ws["A2"].font = bold
    ws["A3"] = subtitle

    # Headers
    header_row = 5
    ws.cell(header_row, 1).value = ""  # labels column
    ws.cell(header_row, 1).font = bold

    periods = pnl["periods"]
    n = len(periods)

    for i, p in enumerate(periods, start=2):
        c = ws.cell(header_row, i)
        c.value = p["label"]
        c.font = bold
        c.alignment = Alignment(horizontal="center")

    total_col = 2 + n
    ws.cell(header_row, total_col).value = "Total"
    ws.cell(header_row, total_col).font = bold
    ws.cell(header_row, total_col).alignment = Alignment(horizontal="center")

    # Helpers
    def write_amount(r: int, c: int, val: float, is_bold: bool = False):
        cell = ws.cell(r, c)
        cell.value = float(val or 0.0)
        cell.number_format = money_fmt
        if is_bold:
            cell.font = bold

    r = header_row + 1

    # Sections
    for sec in pnl["sections"]:
        ws.cell(r, 1).value = sec["section"]
        ws.cell(r, 1).font = bold
        r += 1

        for row in sec["rows"]:
            level = int(row["level"])
            kind = row["kind"]
            label = row["label"]
            vec = row["periods"]
            row_total = row["row_total"]

            is_total = (kind == "total")
            is_group = (kind == "group")

            label_cell = ws.cell(r, 1)
            label_cell.value = label
            label_cell.alignment = Alignment(indent=level * 2)

            if is_total or is_group:
                label_cell.font = bold

            for i, v in enumerate(vec, start=2):
                write_amount(r, i, v, is_bold=is_total)

            write_amount(r, total_col, row_total, is_bold=is_total)
            r += 1

        # Total section line
        ws.cell(r, 1).value = f"Total {sec['section']}"
        ws.cell(r, 1).font = bold
        for i, v in enumerate(sec["section_totals"], start=2):
            write_amount(r, i, v, is_bold=True)
        write_amount(r, total_col, sec["section_total"], is_bold=True)
        r += 2

    # Summary lines
    def write_summary(label: str, vec: list[float]):
        nonlocal r
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = bold
        for i, v in enumerate(vec, start=2):
            write_amount(r, i, v, is_bold=True)
        write_amount(r, total_col, sum(vec), is_bold=True)
        r += 1

    write_summary("Gross Profit", pnl["totals"]["gross_profit"])
    write_summary("Net Earnings", pnl["totals"]["net_profit"])

    # Footer
    r += 2
    ws.cell(r, 1).value = datetime.now().strftime("%A, %b. %d, %Y %I:%M:%S %p") + " - Accruals Basis"

    # Column widths
    ws.column_dimensions["A"].width = 50
    for c in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
