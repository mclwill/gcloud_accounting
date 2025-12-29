# FlaskApp/app/services/balance_sheet_excel.py
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

AUD_ACCOUNTING_FMT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'


def build_balance_sheet_workbook(bs: Dict[str, Any], entity_name: str, subtitle: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    bold = Font(bold=True)
    money_fmt = AUD_ACCOUNTING_FMT

    ws["A1"] = entity_name
    ws["A1"].font = bold
    ws["A2"] = "Balance Sheet"
    ws["A2"].font = bold
    ws["A3"] = subtitle

    cols = bs.get("as_of", [])
    labels = [c.get("label", "") for c in cols]
    n = len(labels)

    header_row = 5

    for i, lbl in enumerate(labels, start=2):
        cell = ws.cell(header_row, i)
        cell.value = lbl
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    def write_amount(r: int, c: int, val: float, is_bold: bool = False):
        cell = ws.cell(r, c)
        cell.value = float(val or 0.0)
        cell.number_format = money_fmt
        if is_bold:
            cell.font = bold

    r = header_row + 1

    for sec in bs.get("sections", []):
        sec_name = sec.get("section", "")
        ws.cell(r, 1).value = sec_name
        ws.cell(r, 1).font = bold
        r += 1

        for row in sec.get("rows", []):
            level = int(row.get("level", 0))
            kind = row.get("kind", "account")
            label = row.get("label", "")
            vec = row.get("cols", [0.0] * n)

            is_total = kind == "total"
            is_group = kind == "group"

            label_cell = ws.cell(r, 1)
            label_cell.value = label
            label_cell.alignment = Alignment(indent=level * 2)
            if is_total or is_group:
                label_cell.font = bold

            for i, v in enumerate(vec, start=2):
                write_amount(r, i, v, is_bold=is_total)

            r += 1

        # Total section (vertical totals are correct)
        ws.cell(r, 1).value = f"Total {sec_name}"
        ws.cell(r, 1).font = bold
        for i, v in enumerate(sec.get("section_totals", [0.0] * n), start=2):
            write_amount(r, i, v, is_bold=True)
        r += 2

    # Accounting equation check (optional, still valid per column)
    diff = bs.get("totals", {}).get("difference", [0.0] * n)
    ws.cell(r, 1).value = "Assets - (Liabilities + Equity)"
    ws.cell(r, 1).font = bold
    for i, v in enumerate(diff, start=2):
        write_amount(r, i, v, is_bold=True)

    r += 2
    ws.cell(r, 1).value = datetime.now().strftime(
        "%A, %b. %d, %Y %I:%M:%S %p"
    ) + " - Accruals Basis"

    ws.column_dimensions["A"].width = 50
    for c in range(2, 2 + n):
        ws.column_dimensions[get_column_letter(c)].width = 18

    return wb



def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
